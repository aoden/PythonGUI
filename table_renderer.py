from openpyxl import *

from tkintertable.Tables import *

from tkintertable.TableModels import *

from dialog import *
from  action_details import ActionDetails

# This class do the table rendering, also contains excel objects, stacks for undo/redo
class TableRenderer:
    def render_table(self, gui_sheet_name='Gui 3', excel_file_name='new exel (1).xlsx', table_area=None, root=None,
                     new=False):
        global k
        data = []
        # declare undo/redo stacks
        self.undo_stack = []
        self.redo_stack = []
        # array contains excel column header, I only add 7 values here, you can add more if you want
        self.col_labels = ['A', 'B', 'C', 'D', 'E', 'F', 'G']
        self.parent = table_area

        if (not new):  # is opening a pre-existed file
            # open excel file
            wb = load_workbook(excel_file_name)
            # get excel sheet
            sh = wb.get_sheet_by_name(gui_sheet_name)
            # assign as instance variable in order to be accessed by other functions
            self.sheet = sh
            self.workbook = wb
            col_num = sh.get_highest_column()
            row_num = sh.get_highest_row()
            # create table model, all tkintertable data can be manipulated by this
            table_model = TableModel()
            # table_model.importDict(data)
            self.table = TableCanvas(table_area, model=table_model, editable=FALSE)
            self.table.createTableFrame()
        else:  # make a new one
            wb = Workbook()
            sh = wb.create_sheet(0, "sheet1")
            # assign as instance variable in order to be accessed by other functions
            self.sheet = sh
            self.workbook = wb
            col_num = 3
            row_num = 10
            table_model = TableModel()
            # table_model.importDict(data)
            # assign as instance variable in order to be accessed by other functions
            self.table = TableCanvas(table_area, model=table_model, editable=FALSE)
            self.table.createTableFrame()
        # bind mouse click event
        root.bind('<ButtonPress-1>', self.clicked)
        # bind double click event
        root.bind('<Double-Button-1>', self.dbclicked)

        for j in range(0, col_num):
            self.table.addColumn(self.col_labels[j])

        for i in range(1, row_num + 1):
            print('enter row ' + str(i))
            for j in range(0, col_num):
                cell = sh[str(self.col_labels[j]) + str(i)].value
                if cell == None:
                    cell = ''
                data.append(cell)
                # print(cell)
                print(data)
            self.table.addRow()
            for k in range(0, col_num):
                self.table.model.data[i][self.col_labels[k]] = data[k]
            data = []
            self.table.redrawTable()

        self.change_emails()

    def clicked(self, event):
        try:
            rclicked = self.table.get_row_clicked(event)
            cclicked = self.table.get_col_clicked(event)
            clicks = (rclicked, cclicked)
            print 'clicks:', clicks
        except:
            print 'Error'
        if clicks:
            # Now we try to get the value of the row+col that was clicked.
            try:
                value = self.table.model.getValueAt(clicks[0], clicks[1])
                # check cell with value is "" or " "
                if (value == '" "' or value == '""'):
                    self.execute_action(clicks)
            except:
                print 'No record at:', clicks

    # this method handle common executions of  dbclicked() and clicked()
    def execute_action(self, clicks):

        # maintain old value in order to undo
        old_val = self.table.model.getValueAt(clicks[0], clicks[1])
        d = MyDialog(self.parent, "enter value")
        self.edit_cell(d.value, clicks[0], clicks[1])
        self.table.redrawTable()
        action_details = ActionDetails(clicks[1], clicks[0], old_val)
        self.undo_stack.insert(0, action_details)

    # change emails values
    def change_emails(self):

        model = self.table.model
        # loop through cells to find emails if any
        for i in range(0, model.getColumnCount() - 1):
            for j in range(0, model.getRowCount() - 1):
                if (re.match("[^@]+@[^@]+\.[^@]+", model.getValueAt(j, i))):  # find email by regular expression
                    print("found " + model.getValueAt(j, i))
                    self.edit_cell('InsDataOps@lexisNexis.com', j, i)  # change email value
                    self.table.redrawTable()

    # handle double click event
    def dbclicked(self, event):

        rclicked = self.table.get_row_clicked(event)
        cclicked = self.table.get_col_clicked(event)
        clicks = (rclicked, cclicked)
        self.execute_action(clicks)
        return

    # this method change tkintertable cell's value
    def edit_cell(self, value, row, col):

        print('edit cell (' + row.__str__() + ',' + col.__str__() + ')')
        self.table.model.setValueAt(value, int(row), int(col))
        self.table.redrawTable()
        return

    # handle undo action
    def undo(self):

        if (self.undo_stack.__len__() != 0):
            action_details = self.undo_stack.pop(0)
            row = action_details.row
            col = action_details.col
            current_val = self.table.model.getValueAt(row, col)

            self.edit_cell(action_details.value, row, col)
            self.redo_stack.insert(0, ActionDetails(row, col, current_val))
        return

    # handle redo action
    def redo(self):

        if (self.redo_stack.__len__() != 0):
            action_details = self.redo_stack.pop(0)
            self.edit_cell(action_details.value, action_details.col, action_details.row)
        return
